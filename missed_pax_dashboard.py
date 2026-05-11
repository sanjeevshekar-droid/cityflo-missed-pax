#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Missed Passenger Dashboard
==========================
Channels : Ticket (source=1) | IO (source=8) | Sage (source=9)
Reasons  : Bus Left Early | Bus Did Not Stop | Bus Took Wrong Route
Extras   : Customer interaction summary per ticket

Usage:
    python missed_pax_dashboard.py                        # May 4 to today
    python missed_pax_dashboard.py --from 2026-05-01
    python missed_pax_dashboard.py --from 2026-05-01 --to 2026-05-08
    python missed_pax_dashboard.py --watch 5             # auto-refresh every 5 min
"""

import sys, io, os, json, re, webbrowser, argparse, time, threading, math
import psycopg2
from datetime import datetime, timezone, timedelta, date
from collections import defaultdict
from html import escape as _h


def haversine_m(lat1, lng1, lat2, lng2):
    R = 6371000
    p = math.pi / 180
    a = (math.sin((lat2 - lat1) * p / 2) ** 2
         + math.cos(lat1 * p) * math.cos(lat2 * p) * math.sin((lng2 - lng1) * p / 2) ** 2)
    return round(2 * R * math.asin(math.sqrt(max(0, a))))

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_FILE = os.path.join(BASE_DIR, 'env')
OUT_FILE = os.path.join(BASE_DIR, 'missed_pax_dashboard.html')

IST = timezone(timedelta(hours=5, minutes=30))

# Primary city lookup: geo_state.id → city name
# Maharashtra=1, Delhi=34, Haryana=35(NCR), UP=36(NCR), Telangana=67, West Bengal=133
STATE_CITY = {1: 'Mumbai', 34: 'Delhi', 35: 'Delhi', 36: 'Delhi',
              67: 'Hyderabad', 133: 'Kolkata'}

# Fallback: bus number letter prefix → city (for tickets with no ride join)
BUS_CITY    = {'C': 'Mumbai', 'D': 'Delhi', 'K': 'Kolkata', 'H': 'Hyderabad'}
SRC_CHANNEL  = {'1': 'Ticket', '8': 'Ticket', '9': 'Sage'}
STATUS_LABEL = {'1': 'Open', '2': 'Resolved'}

# ── Tickets manually excluded from the dashboard ─────────────────────────────
EXCLUDED_TICKET_IDS = {
    12988811,   # Unknown bus + city, no ride data
}

# ── Sage subcategory slug → reason (primary, most accurate) ──────────────────
SLUG_TO_REASON = {
    'the-bus-left-before-the-scheduled-time-without': 'Bus Left Early',
    'the-bus-left-before-the-scheduled-time-wait':    'Bus Left Early',
    'the-driver-didnt-stop-at-my-stop':               'Bus Did Not Stop',
    'didnt-stop-at-designated-stop':                  'Bus Did Not Stop',
}

# Sage category slugs that indicate a missed bus complaint
MISSED_CAT_SLUGS = {
    'my-bus-left-without-me',
    'i-missed-my-bus',
    'driver-related-issue',
    'please-ask-the-driver-to-wait',
    'i-have-an-issue-with-the-driver',
}

# ── Keyword fallback for Ticket / IO channels ────────────────────────────
REASON_KEYWORDS = {
    'Bus Left Early': [
        'bus left early', 'left early', 'bus already left', 'bus departed early',
        'bus left before time', 'bus left before the scheduled', 'left before time',
        'bus came early', 'arrived early and left', 'before i could board',
        'bus was early', 'bus left before', 'bus has already left',
    ],
    'Bus Did Not Stop': [
        'bus did not stop', "bus didn't stop", 'bus did not halt', "bus didn't halt",
        'bus passed without stopping', 'bus skipped my stop', 'bus passed my stop',
        'did not stop at my stop', 'bus drove past', 'skipped the stop',
        'bus went without stopping', 'bus did not pick', "bus didn't pick",
        'did not come to stop', 'didnt stop', 'did not stop', 'bus not stopped',
        'halted ahead', 'didnt halt', 'did not halt',
    ],
}

# Noise phrases in Sage yellow text — skip these
SAGE_NOISE = {
    'how can i help', 'please choose', 'select an option', 'go back',
    'main menu', 'yes', 'no', 'submit', 'okay', 'ok', 'done',
    'i understand', 'thank you', 'please wait', 'connect me', 'other',
}


# ── helpers ──────────────────────────────────────────────────────────────────

def load_env():
    env = {}
    try:
        with open(ENV_FILE, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, _, v = line.partition('=')
                    env[k.strip()] = v.strip()
    except Exception as e:
        print(f'WARN: {e}')
    return env


def aware(t):
    return t if (t and t.tzinfo) else (t.replace(tzinfo=timezone.utc) if t else None)


AGENT_PHRASES = [
    'good morning', 'good afternoon', 'good evening', 'we sincerely apologize',
    'we apologize', 'i apologize', 'apologies for', 'we are sorry', 'inconvenience',
    'relevant team', 'highlighted',
    'we have strictly warned', 'we have warned', 'we will look into',
    'kindly share', 'could you please share', 'please share', 'please provide',
    'thank you for bringing', 'we understand your concern',
]

def is_agent_msg(text):
    if not text: return False
    t = text.lower()
    return any(p in t for p in AGENT_PHRASES)


CONFIRMED_MISS_PHRASES = [
    'as per plot location',          # ops reviewed plot location
    'checked with plot location',    # ops checked
    'checked plot location',
    'plot location',                 # any mention of plot location in ops note
    'as per gps',
    'didn\'t stop at correct',
    'did not stop at correct',
    'not at correct location',
    'not stopped at correct',
    'halted ahead',
    'stopped ahead',
    'halted few meters ahead',
    'cab stop little ahead',
    'confirmed miss', 'genuine miss',
    'warned the driver',
    'warn and guide', 'wran and guide',
    'will make fine', 'made fine', 'fined the driver', 'fine imposed',
    'driver has been warned', 'strictly warned',
    'driver counselled', 'driver counseled',
    'bus left early', 'left before schedule',
    'penalty',
]

CONFIRMED_MISS_EXCLUDE = [
    'here is the plot location link',   # just sharing link, not a confirmation
    'bus was halted on stop',           # bus DID stop — contradicts DNS
    'bus was at stop',
    'bus did stop',
    'did not pick up call',             # cx not answering
    'cx is not picking',
    'customer is not picking',
]

def is_confirmed_miss(text):
    if not text: return False
    t = text.lower()
    if any(ex in t for ex in CONFIRMED_MISS_EXCLUDE):
        return False
    return any(p in t for p in CONFIRMED_MISS_PHRASES)


def get_yellow_items(cmt_text):
    """Extract ordered list of customer-selected yellow-highlighted items from Sage JSON."""
    try:
        data = json.loads(cmt_text)
        items = []
        def walk(obj, in_y=False):
            if isinstance(obj, dict):
                y = obj.get('background', '') == '#FFEEC0' or in_y
                if obj.get('type') == 'Text' and obj.get('value') and y:
                    val = obj['value'].strip()
                    if val: items.append(val)
                for v in obj.values():
                    if isinstance(v, (dict, list)): walk(v, y)
            elif isinstance(obj, list):
                for item in obj: walk(item, in_y)
        walk(data)
        return items
    except Exception:
        return []


def extract_bus_number(text):
    if not text: return None
    m = re.findall(r'"value"\s*:\s*"([CDKH]\d+)"', text, re.IGNORECASE)
    if m: return m[0].upper()
    m = re.findall(r'\b([CDKH]\d{2,4})\b', text, re.IGNORECASE)
    if m: return m[0].upper()
    return None


def detect_reason_keywords(text):
    t = text.lower()
    for reason, patterns in REASON_KEYWORDS.items():
        for p in patterns:
            if p in t:
                return reason
    return None


def build_sage_summary(comments):
    """Build a readable customer journey string from Sage JSON comments."""
    selections = []
    free_text  = []
    for cmt, is_int, *_ in comments:
        if is_int: continue
        items = get_yellow_items(cmt)
        for item in items:
            clean = item.strip()
            if len(clean) < 3: continue
            low = clean.lower()
            if any(noise in low for noise in SAGE_NOISE): continue
            # Check if it looks like free text (sentence) vs menu option
            if len(clean) > 40 or '.' in clean or ',' in clean:
                free_text.append(clean[:180])
            else:
                if clean not in selections:
                    selections.append(clean)
    parts = []
    if selections: parts.append(' → '.join(selections[:6]))
    if free_text:  parts.append(f'"{free_text[0][:150]}"')
    return ' | '.join(parts) if parts else '—'


def build_ticket_summary(comments):
    """Extract first meaningful customer message from plain-text ticket comments."""
    for cmt, is_int, *_ in comments:
        if is_int: continue
        if not cmt or len(cmt.strip()) < 8: continue
        try:
            json.loads(cmt)
            continue  # skip JSON (Sage messages in non-Sage ticket is rare)
        except Exception:
            pass
        if is_agent_msg(cmt): continue
        clean = re.sub(r'\s+', ' ', cmt.strip())
        return clean[:220]
    # If all public comments are agent replies, grab first internal note
    for cmt, is_int, *_ in comments:
        if not is_int: continue
        if not cmt or len(cmt.strip()) < 8: continue
        try: json.loads(cmt); continue
        except: pass
        clean = re.sub(r'\s+', ' ', cmt.strip())
        return f'[Agent note] {clean[:180]}'
    return '—'


def build_ops_comments(comments):
    """Split internal comments by author dept: ops (dept=1) vs CS agent (dept=2).
    Returns (ops_notes, cs_notes) — each a list of {text, name} dicts.
    """
    ops_notes = []
    cs_notes  = []
    for cmt, is_int, _, dept, name in comments:
        if not is_int: continue
        if not cmt or len(cmt.strip()) < 6: continue
        try:
            json.loads(cmt)
            continue  # skip JSON blobs
        except Exception:
            pass
        if is_agent_msg(cmt): continue  # skip template apology messages
        clean = re.sub(r'\s+', ' ', cmt.strip())
        entry = {'text': clean[:220], 'name': (name or '').strip()}
        if str(dept) == '1':  # Operations team
            if not any(e['text'] == entry['text'] for e in ops_notes):
                ops_notes.append(entry)
        elif str(dept) == '2':  # Customer Support
            if not any(e['text'] == entry['text'] for e in cs_notes):
                cs_notes.append(entry)
    return ops_notes[:4], cs_notes[:4]


# ── DB fetch ──────────────────────────────────────────────────────────────────

def load_excel_supplement(excel_path, start_date):
    """
    Read the OPS-maintained missed-pax Excel form and return a dict of
    {ticket_id: reason} for rows from start_date onwards.
    Only rows with a valid numeric ticket ID are included.
    Deduplication: first occurrence per ticket_id wins.
    """
    try:
        import openpyxl
    except ImportError:
        return {}
    if not os.path.exists(excel_path):
        return {}

    LAPSE_MAP = {
        'bus left early':  'Bus Left Early',
        'left early':      'Bus Left Early',
        'bus not halted':  'Bus Did Not Stop',
        'bus not stopped': 'Bus Did Not Stop',
        'did not stop':    'Bus Did Not Stop',
        'didnt stop':      'Bus Did Not Stop',
        'not halted':      'Bus Did Not Stop',
    }

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    result = {}
    for r in ws.iter_rows(values_only=True):
        ts = r[0]
        if not isinstance(ts, datetime): continue
        if ts.date() < start_date: continue
        raw_tid = r[10]
        if not raw_tid: continue
        try:
            tid = int(float(raw_tid))
        except Exception:
            continue
        if tid > 99_999_999:   # reject phone numbers accidentally entered
            continue
        if tid in result:
            continue           # dedup: first occurrence wins
        lapse = str(r[5]).strip().lower() if r[5] else ''
        reason = next((v for k, v in LAPSE_MAP.items() if k in lapse), None)
        if reason:
            result[tid] = reason
    wb.close()
    return result


def fetch_data(db_url, start_utc, end_utc, excel_supplement=None):
    conn = psycopg2.connect(db_url)
    cur  = conn.cursor()

    # 1. Sage tickets with relevant slugs (primary — most accurate)
    slug_list = list(SLUG_TO_REASON.keys()) + ['i-was-at-the-stop-but-couldnt-find-the-bus']
    ph_slugs  = ','.join(['%s'] * len(slug_list))
    cur.execute(f"""
        SELECT t.id, t.source, t.created, t.sageai_category_slug, t.sageai_subcategory_slug, t.status
        FROM support_ticket t
        WHERE t.created >= %s AND t.created <= %s
          AND t.source = '9'
          AND t.sageai_subcategory_slug IN ({ph_slugs})
        ORDER BY t.id
    """, (start_utc, end_utc, *slug_list))
    sage_slug_rows = cur.fetchall()

    # 2. Non-Sage tickets (source=1, source=8) — two-step to avoid timeout:
    #    Step A: get all ticket IDs in date range for these sources
    cur.execute("""
        SELECT id, source, created, sageai_category_slug, sageai_subcategory_slug, status
        FROM support_ticket
        WHERE created >= %s AND created <= %s
          AND source IN ('1','8')
        ORDER BY id
    """, (start_utc, end_utc))
    non_sage_tickets = cur.fetchall()
    non_sage_ids = [r[0] for r in non_sage_tickets]
    non_sage_meta = {r[0]: r for r in non_sage_tickets}

    ticket_rows = []
    if non_sage_ids:
        #    Step B: scan comments only for those ticket IDs
        ph_ids = ','.join(['%s'] * len(non_sage_ids))
        cur.execute(f"""
            SELECT DISTINCT ticket_id
            FROM support_ticketcomment
            WHERE ticket_id IN ({ph_ids})
              AND (
                comment ILIKE '%%bus left early%%'
                OR comment ILIKE '%%left early%%'
                OR comment ILIKE '%%bus already left%%'
                OR comment ILIKE '%%bus did not stop%%'
                OR comment ILIKE '%%bus didn''t stop%%'
                OR comment ILIKE '%%did not stop%%'
                OR comment ILIKE '%%didnt stop%%'
                OR comment ILIKE '%%did not halt%%'
                OR comment ILIKE '%%didnt halt%%'
                OR comment ILIKE '%%bus skipped%%'
                OR comment ILIKE '%%bus passed without%%'
                OR comment ILIKE '%%bus passed my stop%%'
                OR comment ILIKE '%%halted ahead%%'
              )
        """, non_sage_ids)
        matched_ids = [r[0] for r in cur.fetchall()]
        ticket_rows = [non_sage_meta[tid] for tid in matched_ids if tid in non_sage_meta]

    all_rows  = sage_slug_rows + ticket_rows
    all_ids   = [r[0] for r in all_rows]
    ticket_meta = {r[0]: {'source': str(r[1]), 'created': r[2],
                           'cat': r[3], 'subcat': r[4], 'status': str(r[5] or '')} for r in all_rows}

    # Inject Excel supplement: tickets logged by OPS that the keyword/slug filter missed
    if excel_supplement:
        existing = set(all_ids)
        excel_new = [tid for tid in excel_supplement if tid not in existing]
        if excel_new:
            ph_ex = ','.join(['%s'] * len(excel_new))
            cur.execute(f"""
                SELECT id, source, created, sageai_category_slug, sageai_subcategory_slug, status
                FROM support_ticket
                WHERE id IN ({ph_ex})
                  AND created >= %s AND created <= %s
            """, (*excel_new, start_utc, end_utc))
            for row in cur.fetchall():
                tid = row[0]
                all_ids.append(tid)
                ticket_meta[tid] = {'source': str(row[1]), 'created': row[2],
                                    'cat': row[3], 'subcat': row[4], 'status': str(row[5] or '')}

    if not all_ids:
        conn.close()
        return []

    # Fetch bus number + city + boarding + GPS keys + assignee names for all tickets at once
    ph2 = ','.join(['%s'] * len(all_ids))
    cur.execute(f"""
        SELECT t.id,
               COALESCE(r.bus_number, '')           AS bus_num,
               gs.id                                AS state_id,
               r.boarding_status,
               r.vehicle_ride_id,
               r.start_stop_id,
               TRIM(COALESCE(p1.first_name,'') || ' ' || COALESCE(p1.last_name,'')) AS assigned_name,
               TRIM(COALESCE(p2.first_name,'') || ' ' || COALESCE(p2.last_name,'')) AS ops_name,
               gl.lat                               AS stop_lat,
               gl.lng                               AS stop_lng,
               rs.name                              AS stop_name
        FROM support_ticket t
        LEFT JOIN rides_customerride r  ON r.id  = t.customer_ride_id
        LEFT JOIN routes_stop       rs  ON rs.id = r.start_stop_id
        LEFT JOIN geo_location      gl  ON gl.id = rs.location_id
        LEFT JOIN geo_state         gs  ON gs.id = rs.state_id
        LEFT JOIN users_employee    e1  ON e1.id = t.assigned_to_employee_id
        LEFT JOIN users_person      p1  ON p1.id = e1.person_id
        LEFT JOIN users_employee    e2  ON e2.id = t.collaborator_employee_id
        LEFT JOIN users_person      p2  ON p2.id = e2.person_id
        WHERE t.id IN ({ph2})
    """, all_ids)
    ride_info = {row[0]: (row[1] or '', row[2], row[3], row[4], row[5], row[6] or '', row[7] or '',
                          row[8], row[9], row[10] or '')
                 for row in cur.fetchall()}
    # ride_info[tid] = (bus_num, state_id, boarding_status, vehicle_ride_id, start_stop_id,
    #                   assigned_name, ops_name, stop_lat, stop_lng, stop_name)

    # GPS stop records: fetch both arrival (trigger='1') and departure (trigger='2')
    # Used for: BLE → departure delay; DNS → dwell time at stop
    all_vrids = list(set(v[3] for v in ride_info.values() if v[3]))
    all_sids  = list(set(v[4] for v in ride_info.values() if v[4]))
    dep_delay_map = {}   # (vehicle_ride_id, start_stop_id) → delay_seconds (departure)
    gps_map       = {}   # (vehicle_ride_id, start_stop_id, trigger_type) → (ts, lat, lng)
    if all_vrids and all_sids:
        ph_v = ','.join(['%s'] * len(all_vrids))
        ph_s = ','.join(['%s'] * len(all_sids))
        cur.execute(f"""
            SELECT vehicle_ride_id, stop_id, trigger_type, delay, location_timestamp, lat, lng
            FROM vehicles_vehicleridestoplocationlog
            WHERE vehicle_ride_id IN ({ph_v})
              AND stop_id IN ({ph_s})
        """, (*all_vrids, *all_sids))
        for vrid, sid, ttype, delay, ts, lat, lng in cur.fetchall():
            gps_map[(vrid, sid, ttype)] = (ts, lat, lng)
            if ttype == '2':
                dep_delay_map[(vrid, sid)] = delay

    # Tier-2 city fallback: for tickets with no ride-based state, use customer's latest ride city
    no_state_ids = [tid for tid, info in ride_info.items() if not info[1]]
    cid_state    = {}   # customer_id → state_id
    ticket_cid   = {}   # ticket_id  → customer_id  (for no-state tickets only)
    if no_state_ids:
        ph_ns = ','.join(['%s'] * len(no_state_ids))
        cur.execute(f"SELECT id, customer_id FROM support_ticket WHERE id IN ({ph_ns})", no_state_ids)
        ticket_cid = {row[0]: row[1] for row in cur.fetchall() if row[1]}
        cids = list(set(ticket_cid.values()))
        if cids:
            ph_c = ','.join(['%s'] * len(cids))
            cur.execute(f"""
                SELECT DISTINCT ON (r.customer_id)
                    r.customer_id, gs.id
                FROM rides_customerride r
                LEFT JOIN routes_stop rs ON rs.id = r.start_stop_id
                LEFT JOIN geo_state   gs ON gs.id = rs.state_id
                WHERE r.customer_id IN ({ph_c})
                  AND r.bus_number IS NOT NULL
                  AND gs.id IS NOT NULL
                ORDER BY r.customer_id, r.created DESC
            """, cids)
            cid_state = {row[0]: row[1] for row in cur.fetchall()}

    # Fetch all comments for these tickets (include author_id for dept lookup)
    cur.execute(f"""
        SELECT ticket_id, comment, is_internal, created, author_id
        FROM support_ticketcomment
        WHERE ticket_id IN ({ph2})
        ORDER BY ticket_id, created ASC
    """, all_ids)
    raw_cmts = cur.fetchall()

    # Build author info map (author_id = users_person.user_id)
    # dept=1 → Operations team; dept=2 → Customer Support
    all_author_ids = list(set(row[4] for row in raw_cmts if row[4]))
    author_info = {}  # user_id → {'dept': int, 'name': str}
    if all_author_ids:
        ph_auth = ','.join(['%s'] * len(all_author_ids))
        cur.execute(f"""
            SELECT up.user_id, ue.department, up.first_name, up.last_name
            FROM users_person up
            JOIN users_employee ue ON ue.person_id = up.id
            WHERE up.user_id IN ({ph_auth})
        """, all_author_ids)
        for uid, dept, fn, ln in cur.fetchall():
            author_info[uid] = {'dept': dept, 'name': ((fn or '') + ' ' + (ln or '')).strip()}

    conn.close()

    by_ticket = defaultdict(list)
    for tid, cmt, is_int, created, author_id in raw_cmts:
        info = author_info.get(author_id, {})
        by_ticket[tid].append((cmt or '', bool(is_int), created, info.get('dept'), info.get('name', '')))

    results = []
    for tid in all_ids:
        if tid in EXCLUDED_TICKET_IDS:
            continue
        meta     = ticket_meta[tid]
        source   = meta['source']
        channel  = SRC_CHANNEL.get(source, f'src-{source}')
        created  = aware(meta['created']).astimezone(IST)
        subcat   = meta.get('subcat') or ''
        comments = by_ticket.get(tid, [])

        # Reason detection: slug first, then keyword fallback
        reason = SLUG_TO_REASON.get(subcat)
        if not reason:
            all_text  = []
            for cmt, *_ in comments:
                if not cmt: continue
                try:
                    items = get_yellow_items(cmt)
                    all_text.extend(items)
                except Exception:
                    all_text.append(cmt)
            combined = ' '.join(all_text)
            reason = detect_reason_keywords(combined)
            if not reason:
                raw_all = ' '.join(c for c, *_ in comments)
                reason  = detect_reason_keywords(raw_all)

        if not reason:
            reason = (excel_supplement or {}).get(tid)
        if not reason:
            continue  # skip if can't categorise

        # BLE validation: use GPS departure timing + boarding status
        info = ride_info.get(tid, ('', None, None, None, None, '', '', None, None, ''))
        if reason == 'Bus Left Early':
            boarding_status = info[2]
            vrid = info[3]; sid = info[4]
            dep_delay = dep_delay_map.get((vrid, sid)) if (vrid and sid) else None

            if dep_delay is not None:
                if dep_delay >= 0:
                    continue  # GPS: bus on time or late — not a genuine BLE
            else:
                if boarding_status == 'success':
                    continue  # no GPS + system says boarded = insufficient evidence

        # Bus number — ride join is primary, comment extraction is fallback
        ride_bus_num  = info[0]
        state_id      = info[1]
        assigned_name = info[5] or '—'
        ops_name      = info[6] or '—'

        bus_num = ride_bus_num.strip().upper() if ride_bus_num else ''
        if not bus_num:
            all_raw = ' '.join(c for c, *_ in comments)
            bus_num = (extract_bus_number(all_raw) or '').upper()
        bus_num = bus_num or 'Unknown'

        # City — 3-tier resolution:
        # 1. State from this ticket's linked ride stop (most accurate)
        # 2. State from customer's latest historical ride (for tickets with no ride_id)
        # 3. Bus number prefix fallback (C/D/K/H)
        city = STATE_CITY.get(state_id) if state_id else None
        if not city:
            cid      = ticket_cid.get(tid)
            fallback = STATE_CITY.get(cid_state.get(cid)) if cid else None
            city     = fallback or BUS_CITY.get(bus_num[:1], 'Unknown')

        # Customer interaction summary
        if channel == 'Sage':
            summary = build_sage_summary(comments)
        else:
            summary = build_ticket_summary(comments)

        ops_notes, cs_notes = build_ops_comments(comments)
        status_raw = meta.get('status', '')
        status = STATUS_LABEL.get(status_raw, 'Open')

        # Confirmed miss: ops internal comment explicitly validates the complaint
        confirmed = any(
            is_confirmed_miss(cmt)
            for cmt, is_int, *_ in comments
            if is_int and cmt and not is_agent_msg(cmt)
        )

        # GPS plot data
        vrid = info[3]; sid = info[4]
        stop_lat = info[7]; stop_lng = info[8]; stop_name = info[9]
        plot = None
        if vrid and sid:
            arr = gps_map.get((vrid, sid, '1'))
            dep = gps_map.get((vrid, sid, '2'))
            if arr or stop_lat:
                dwell = int((dep[0] - arr[0]).total_seconds()) if (arr and dep) else None
                arr_dist = haversine_m(arr[1], arr[2], stop_lat, stop_lng) if (arr and stop_lat) else None
                dep_dist = haversine_m(dep[1], dep[2], stop_lat, stop_lng) if (dep and stop_lat) else None
                # Closest the bus GPS came to the stop
                min_dist = min(d for d in [arr_dist, dep_dist] if d is not None) if (arr_dist or dep_dist) else None
                plot = {
                    'stop_lat':  stop_lat,  'stop_lng':  stop_lng,
                    'stop_name': stop_name or '',
                    'arr_lat':   arr[1] if arr else None,
                    'arr_lng':   arr[2] if arr else None,
                    'arr_ts':    aware(arr[0]).astimezone(IST).strftime('%H:%M:%S') if arr else None,
                    'dep_lat':   dep[1] if dep else None,
                    'dep_lng':   dep[2] if dep else None,
                    'dep_ts':    aware(dep[0]).astimezone(IST).strftime('%H:%M:%S') if dep else None,
                    'dwell_sec': dwell,
                    'arr_dist':  arr_dist,
                    'dep_dist':  dep_dist,
                    'min_dist':  min_dist,
                }

        results.append({
            'ticket_id':    tid,
            'channel':      channel,
            'date_str':     created.strftime('%d %b %Y'),
            'date_iso':     created.strftime('%Y-%m-%d'),
            'bus_num':      bus_num or 'Unknown',
            'city':         city,
            'reason':       reason,
            'summary':      summary,
            'subcat':       subcat,
            'assigned':     assigned_name,
            'ops_assignee': ops_name,
            'status':       status,
            'ops_notes':    ops_notes,
            'cs_notes':     cs_notes,
            'confirmed':    confirmed,
            'plot':         plot,
            'gps_halted':   bool(plot and plot.get('dwell_sec') is not None and plot['dwell_sec'] >= 15),
        })

    # Remove DNS records where GPS confirms bus halted ≥15s — not genuine misses
    results = [r for r in results
               if not (r['reason'] == 'Bus Did Not Stop' and r.get('gps_halted'))]

    return results


# ── Python-side table rendering (no-JS fallback) ────────────────────────────

def _py_summary_html(s):
    if not s or s == '—':
        return '<span style="color:#444">—</span>'
    parts = s.split(' | ')
    out = ''
    for part in parts:
        e = _h(part)
        if part.startswith('"') and part.endswith('"'):
            out += f'<div class="quote">{e}</div>'
        else:
            out += f'<div class="flow">{e}</div>'
    return out


def _py_notes_html(notes, team_label, badge_color):
    if not notes:
        return '<span style="color:#333">—</span>'
    parts = []
    for n in notes:
        text = _h(n if isinstance(n, str) else (n.get('text') or ''))
        name = _h('' if isinstance(n, str) else (n.get('name') or ''))
        hdr = (f'<div style="margin-bottom:3px">'
               f'<span style="background:{badge_color};color:#fff;font-size:.62rem;font-weight:700;'
               f'padding:1px 6px;border-radius:3px;letter-spacing:.4px;text-transform:uppercase">'
               f'{team_label}</span>'
               f' <span style="color:#aaa;font-size:.72rem">{name}</span></div>') if name else ''
        parts.append(f'<div class="ops-note">{hdr}{text}</div>')
    return ''.join(parts)


def build_static_rows(rows):
    CITY_B   = {'Mumbai':'b-Mumbai','Delhi':'b-Delhi','Kolkata':'b-Kolkata',
                'Hyderabad':'b-Hyderabad','Unknown':'b-Unknown'}
    REASON_B = {'Bus Did Not Stop':'b-dns','Bus Left Early':'b-ble'}
    CHAN_B   = {'Sage':'b-Sage','Ticket':'b-Ticket'}
    RLBL     = {'Bus Did Not Stop':'Did Not Stop','Bus Left Early':'Left Early'}

    bus_day = defaultdict(int)
    for r in rows:
        if r['bus_num'] != 'Unknown':
            bus_day[f"{r['bus_num']}|{r['date_iso']}"] += 1

    out = []
    for r in rows:
        key       = f"{r['bus_num']}|{r['date_iso']}"
        cnt       = bus_day[key]
        repeat    = cnt > 1
        confirmed = bool(r.get('confirmed'))
        row_cls   = 'repeat-bus' if repeat else ('confirmed-row' if confirmed else '')

        date_cell = _h(r['date_str'])
        if confirmed:
            date_cell += (' <span class="badge b-confirmed" '
                          'style="font-size:.65rem;padding:1px 5px">&#10003; Confirmed</span>')

        bus_cls  = 'bus-tag' + (' repeat-bus-tag' if repeat else '')
        bus_cell = f'<span class="{bus_cls}">{_h(r["bus_num"])}'
        if repeat:
            bus_cell += f' <span class="repeat-badge">{cnt}x</span>'
        bus_cell += '</span>'

        plot_btn = (f'<br><button class="plot-btn" onclick="openPlot({r["ticket_id"]})">'
                    '&#128205; Plot</button>') if r.get('plot') else ''

        # Build search text for keyword filtering (lowercased, stored in data attribute)
        def _notes_text(lst):
            return ' '.join((n if isinstance(n, str) else (n.get('text') or ''))
                            for n in (lst or []))
        search_raw = ' '.join([
            r.get('summary') or '',
            r['bus_num'],
            _notes_text(r.get('cs_notes')),
            _notes_text(r.get('ops_notes')),
        ]).lower()
        search_attr = _h(re.sub(r'\s+', ' ', search_raw).strip())

        # Sort key attributes (used by sortAndRender to reorder DOM nodes)
        sort_summary = _h((r.get('summary') or '').strip())
        sort_assigned = _h(r.get('assigned') or '')
        sort_ops_assigned = _h(r.get('ops_assignee') or '')

        out.append(
            f'<tr class="{row_cls}" data-tid="{r["ticket_id"]}"'
            f' data-date="{r["date_iso"]}" data-bus="{_h(r["bus_num"])}"'
            f' data-city="{_h(r["city"])}" data-reason="{_h(r["reason"])}"'
            f' data-channel="{_h(r["channel"])}" data-status="{_h(r["status"])}"'
            f' data-confirmed="{"true" if confirmed else "false"}"'
            f' data-summary="{sort_summary}" data-assigned="{sort_assigned}"'
            f' data-opsassigned="{sort_ops_assigned}"'
            f' data-search="{search_attr}">'
            f'<td style="white-space:nowrap">{date_cell}</td>'
            f'<td>{bus_cell}</td>'
            f'<td><span class="badge {CITY_B.get(r["city"],"b-Unknown")}">{_h(r["city"])}</span></td>'
            f'<td><span class="badge {REASON_B.get(r["reason"],"")}">'
            f'{RLBL.get(r["reason"], _h(r["reason"]))}</span></td>'
            f'<td><span class="badge {CHAN_B.get(r["channel"],"")}">{_h(r["channel"])}</span></td>'
            f'<td><span class="badge b-{_h(r["status"])}">{_h(r["status"])}</span></td>'
            f'<td class="summary-cell">{_py_summary_html(r.get("summary") or "—")}</td>'
            f'<td class="ops-cell">{_py_notes_html(r.get("cs_notes",[]),"Support","#6b7280")}</td>'
            f'<td class="ops-cell">{_py_notes_html(r.get("ops_notes",[]),"Operations","#7c3aed")}</td>'
            f'<td class="assignee-cell">{_h(r.get("assigned") or "—")}</td>'
            f'<td class="assignee-cell">{_h(r.get("ops_assignee") or "—")}</td>'
            f'<td style="white-space:nowrap"><span class="tid-link">#{r["ticket_id"]}</span>'
            f'{plot_btn}</td>'
            f'</tr>'
        )
    return '\n'.join(out)


# ── HTML ──────────────────────────────────────────────────────────────────────

def build_html(rows, start_label, end_label, generated_at, refresh_min=0):
    if not rows:
        return '<html><body style="background:#0f1117;color:#eee;font-family:sans-serif;padding:40px"><h2>No data found.</h2></body></html>'

    static_rows = build_static_rows(rows)

    all_dates   = sorted(set(r['date_iso'] for r in rows))
    all_reasons = ['Bus Did Not Stop', 'Bus Left Early']
    all_channels= sorted(set(r['channel'] for r in rows))
    all_cities  = [c for c in ['Mumbai', 'Delhi', 'Kolkata', 'Hyderabad', 'Unknown']
                   if any(r['city'] == c for r in rows)]

    total       = len(rows)
    uniq_keys   = set((r['bus_num'], r['date_iso']) for r in rows if r['bus_num'] != 'Unknown')
    unique_trips= len(uniq_keys)

    by_reason   = defaultdict(list)
    for r in rows: by_reason[r['reason']].append(r)
    dns       = len(by_reason.get('Bus Did Not Stop', []))
    ble       = len(by_reason.get('Bus Left Early', []))
    confirmed = sum(1 for r in rows if r.get('confirmed'))

    def pct(n): return round(n / total * 100) if total else 0

    # Charts
    day_data = {d: {r: 0 for r in all_reasons} for d in all_dates}
    for r in rows:
        if r['date_iso'] in day_data: day_data[r['date_iso']][r['reason']] += 1

    chart_labels = json.dumps(all_dates)
    chart_dns = json.dumps([day_data[d]['Bus Did Not Stop'] for d in all_dates])
    chart_ble = json.dumps([day_data[d]['Bus Left Early'] for d in all_dates])

    city_counts = defaultdict(int)
    for r in rows: city_counts[r['city']] += 1
    donut_labels = json.dumps(list(city_counts.keys()))
    donut_data   = json.dumps([city_counts[c] for c in city_counts])
    city_colors_map = {'Mumbai':'#3498db','Delhi':'#e74c3c','Kolkata':'#2ecc71',
                       'Hyderabad':'#f39c12','Unknown':'#636e72'}
    donut_colors = json.dumps([city_colors_map.get(c, '#aaa') for c in city_counts])

    # City-wise breakdown
    city_order = ['Mumbai', 'Delhi', 'Kolkata', 'Hyderabad', 'Unknown']
    city_stats = {}
    for r in rows:
        c = r['city']
        if c not in city_stats:
            city_stats[c] = {'total': 0, 'uniq': set()}
        city_stats[c]['total'] += 1
        if r['bus_num'] != 'Unknown':
            city_stats[c]['uniq'].add((r['bus_num'], r['date_iso']))

    city_color = {'Mumbai':'#3498db','Delhi':'#e74c3c','Kolkata':'#2ecc71',
                  'Hyderabad':'#f39c12','Unknown':'#636e72'}

    city_cards_html = ''
    for city in city_order:
        if city not in city_stats: continue
        s = city_stats[city]
        col = city_color.get(city, '#aaa')
        city_cards_html += f'''
  <div class="city-card" onclick="filterCity('{city}')" title="Click to filter by {city}">
    <div class="city-card-name" style="border-left:3px solid {col};padding-left:8px">{city}</div>
    <div class="city-card-nums">
      <span class="city-stat"><span class="city-stat-val">{s["total"]}</span><span class="city-stat-lbl">Total</span></span>
      <span class="city-stat"><span class="city-stat-val">{len(s["uniq"])}</span><span class="city-stat-lbl">Unique</span></span>
    </div>
  </div>'''

    # Channel summary pills
    ch_counts = defaultdict(int)
    for r in rows: ch_counts[r['channel']] += 1
    ch_html = ''.join(
        f'<div class="ch-pill ch-{ch.replace("/","").replace(" ","")}"><span class="ch-name">{ch}</span></div>'
        for ch in ['Sage', 'Ticket'] if ch in ch_counts
    )

    # Filter options
    reason_opts    = ''.join(f'<option value="{r}">{r}</option>' for r in all_reasons)
    city_opts      = ''.join(f'<option value="{c}">{c}</option>' for c in all_cities)
    ch_opts        = ''.join(f'<option value="{c}">{c}</option>' for c in ['Sage', 'Ticket'] if c in all_channels)
    # Generate week buttons (Mon–Sun) spanning the full data range
    from datetime import date as _date, timedelta as _td
    _first = _date.fromisoformat(all_dates[0])
    _last  = _date.fromisoformat(all_dates[-1])
    _wks = []
    _mon = _first - _td(days=_first.weekday())
    while _mon <= _last:
        _sun = _mon + _td(days=6)
        _ws  = max(_mon, _first); _we = min(_sun, _last)
        _ws_s = _ws.isoformat(); _we_s = _we.isoformat()
        if _ws.month == _we.month:
            _lbl = f"{_ws.day}–{_we.day} {_ws.strftime('%b')}"
        else:
            _lbl = f"{_ws.day} {_ws.strftime('%b')}–{_we.day} {_we.strftime('%b')}"
        _wks.append((_ws_s, _we_s, _lbl))
        _mon += _td(days=7)
    week_btns_html = ''.join(
        f'<button class="week-btn" data-start="{ws}" data-end="{we}" onclick="selectWeek(this)">{lbl}</button>'
        for ws, we, lbl in _wks
    )

    # Auto-reload support
    if refresh_min:
        refresh_ms = refresh_min * 60 * 1000
        auto_reload_js = f'setTimeout(()=>location.reload(), {refresh_ms});'
        refresh_badge  = (f'<span style="font-size:.7rem;color:#a78bfa;background:#1a1a3a;'
                          f'border:1px solid #7c3aed40;border-radius:10px;padding:2px 8px">'
                          f'&#9679; Auto-refresh every {refresh_min} min</span>')
    else:
        auto_reload_js = ''
        refresh_badge  = ''

    # Replace </ with <\/ so ops notes never accidentally close a <script> tag
    rows_json = json.dumps(rows).replace('</', '<\\/')

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Missed Pax Dashboard — Cityflo CS</title>
<script src="chart.umd.min.js"></script>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js" async></script>
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js" async></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#0d0f18;color:#dde1ec;min-height:100vh}}

/* header */
.topbar{{background:#111420;border-bottom:3px solid #c0392b;padding:14px 28px;
         display:flex;align-items:center;justify-content:space-between}}
.topbar h1{{font-size:1.2rem;font-weight:700;color:#fff;letter-spacing:.3px}}
.topbar h1 em{{color:#e74c3c;font-style:normal}}
.top-meta{{font-size:.76rem;color:#666;text-align:right;line-height:1.7}}
.top-meta strong{{color:#aaa}}

/* kpi */
.kpi-row{{display:flex;gap:12px;padding:18px 28px 0}}
.kpi{{flex:1;background:#151824;border-radius:10px;padding:16px 18px;
      border-left:4px solid #333;transition:.2s;cursor:default}}
.kpi:hover{{transform:translateY(-2px);box-shadow:0 6px 24px rgba(0,0,0,.5)}}
.kpi-val{{font-size:2.1rem;font-weight:800;line-height:1;margin-bottom:4px}}
.kpi-label{{font-size:.73rem;text-transform:uppercase;letter-spacing:.6px;color:#777}}
.kpi-sub{{font-size:.7rem;color:#555;margin-top:3px}}
.k-total{{border-color:#4fc3f7}} .k-total .kpi-val{{color:#4fc3f7}}
.k-uniq {{border-color:#27ae60}} .k-uniq  .kpi-val{{color:#27ae60}}
.k-dns  {{border-color:#e74c3c}} .k-dns   .kpi-val{{color:#e74c3c}}
.k-ble  {{border-color:#f39c12}} .k-ble   .kpi-val{{color:#f39c12}}
.k-confirmed{{border-color:#00b894}} .k-confirmed .kpi-val{{color:#00b894}}
.k-confirmed:hover{{box-shadow:0 0 0 2px #00b89440}}
.confirmed-row{{background:#051a10!important;border-left:3px solid #00b894}}
.confirmed-row:hover{{background:#072415!important}}
.b-confirmed{{background:#052e1a;color:#00b894;border:1px solid #00b89450}}
.conf-section{{padding:0 28px 18px}}
.conf-hdr{{display:flex;align-items:center;gap:10px;padding:12px 0 10px;
           border-top:2px solid #00b89440}}
.conf-hdr-title{{font-size:.8rem;font-weight:700;color:#00b894;text-transform:uppercase;
                  letter-spacing:.5px}}
.conf-hdr-count{{background:#00b89420;color:#00b894;border:1px solid #00b89440;
                  border-radius:10px;padding:1px 9px;font-size:.75rem;font-weight:700}}

/* charts */
.charts-row{{display:flex;gap:12px;padding:16px 28px}}
.chart-card{{background:#151824;border-radius:10px;padding:18px 20px}}
.chart-card.main{{flex:2}}
.chart-card.side{{flex:1}}
.chart-title{{font-size:.76rem;font-weight:600;color:#888;text-transform:uppercase;
              letter-spacing:.5px;margin-bottom:14px}}

/* city breakdown */
.city-row{{padding:6px 28px 16px;display:flex;align-items:flex-start;gap:14px;flex-wrap:wrap}}
.city-row-label{{font-size:.7rem;color:#555;text-transform:uppercase;letter-spacing:.5px;
                 padding-top:10px;white-space:nowrap;min-width:90px}}
.city-cards{{display:flex;gap:10px;flex-wrap:wrap;flex:1}}
.city-card{{background:#111420;border:1px solid #1e2235;border-radius:8px;
            padding:10px 14px;cursor:pointer;transition:border-color .15s;min-width:150px}}
.city-card:hover{{border-color:#4fc3f7}}
.city-card-name{{font-size:.78rem;font-weight:700;color:#ccc;margin-bottom:8px}}
.city-card-nums{{display:flex;gap:10px}}
.city-stat{{display:flex;flex-direction:column;align-items:center;min-width:36px}}
.city-stat-val{{font-size:1rem;font-weight:700;color:#e0e0e0;line-height:1.1}}
.city-stat-lbl{{font-size:.58rem;color:#555;text-transform:uppercase;letter-spacing:.3px}}
.dns-stat .city-stat-val{{color:#e74c3c}}
.ble-stat .city-stat-val{{color:#f39c12}}
.conf-stat .city-stat-val{{color:#00b894}}
/* channel pills */
.ch-strip{{padding:0 28px 14px;display:flex;align-items:center;gap:10px;flex-wrap:wrap}}
.ch-label{{font-size:.72rem;color:#555;text-transform:uppercase;letter-spacing:.5px}}
.ch-pill{{display:flex;align-items:center;gap:8px;border-radius:20px;
          padding:5px 14px;font-size:.8rem;border:1px solid #2a2e40}}
.ch-Sage{{background:#1b1040;border-color:#6c3483}} .ch-Sage .ch-name{{color:#a569bd}}
.ch-Ticket{{background:#0f2a1a;border-color:#1e7a40}} .ch-Ticket .ch-name{{color:#52be80}}
.ch-count{{background:#1e2235;border-radius:10px;padding:1px 9px;
           font-weight:700;color:#fff;font-size:.78rem}}

/* filter bar */
.filter-bar{{background:#111420;border-top:1px solid #1e2235;border-bottom:1px solid #1e2235;
             padding:10px 28px;display:flex;gap:10px;align-items:center;flex-wrap:wrap}}
.filter-bar label{{font-size:.72rem;color:#555;text-transform:uppercase;letter-spacing:.4px}}
.filter-bar select,.filter-bar input{{
  background:#0d0f18;color:#ccc;border:1px solid #1e2235;border-radius:6px;
  padding:5px 9px;font-size:.8rem;outline:none}}
.filter-bar select:focus,.filter-bar input:focus{{border-color:#4fc3f7}}
.filter-bar input{{width:150px}}
.f-count{{margin-left:auto;font-size:.78rem;color:#555}}
.f-count strong{{color:#4fc3f7}}
.btn-reset{{background:#1e2235;color:#999;border:none;border-radius:6px;
            padding:5px 14px;cursor:pointer;font-size:.78rem}}
.btn-reset:hover{{background:#252c45;color:#fff}}

/* section header */
.sec-hdr{{padding:14px 28px 6px;font-size:.72rem;color:#555;
          text-transform:uppercase;letter-spacing:.5px;border-top:1px solid #151824}}

/* table */
.tbl-wrap{{padding:0 28px 24px;overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:.8rem}}
thead th{{background:#111420;color:#777;font-weight:600;text-transform:uppercase;
          letter-spacing:.4px;padding:9px 10px;text-align:left;
          border-bottom:1px solid #1e2235;white-space:nowrap;cursor:pointer;user-select:none}}
thead th:hover{{color:#ccc}}
tbody tr{{border-bottom:1px solid #151824}}
tbody tr:hover{{background:#111420}}
tbody td{{padding:8px 10px;vertical-align:top}}

/* badges */
.badge{{display:inline-block;border-radius:4px;padding:2px 7px;font-size:.72rem;font-weight:600;white-space:nowrap}}
.b-dns {{background:#3d0a0a;color:#e74c3c}}
.b-ble {{background:#3d2200;color:#f39c12}}
.b-Mumbai{{background:#0a2235;color:#5dade2}}
.b-Delhi{{background:#3d0a0a;color:#ec7063}}
.b-Kolkata{{background:#0a2e14;color:#58d68d}}
.b-Hyderabad{{background:#3d2800;color:#f0b27a}}
.b-Unknown{{background:#1a1a1a;color:#888}}
.b-Sage{{background:#1b1040;color:#a569bd}}
.b-Ticket{{background:#0f2a1a;color:#52be80}}
.b-Open{{background:#0a2e14;color:#58d68d}}
.b-Resolved{{background:#1a1a2e;color:#7f8c8d}}

/* week buttons & date range picker */
.week-picker-wrap{{display:flex;align-items:center;gap:5px;flex-wrap:wrap}}
.week-btn{{background:#1a1d2e;color:#888;border:1px solid #1e2235;border-radius:6px;
           padding:4px 10px;font-size:.73rem;cursor:pointer;transition:.15s;white-space:nowrap}}
.week-btn:hover{{border-color:#4fc3f7;color:#ccc}}
.week-btn.active{{background:#0d2a40;border-color:#4fc3f7;color:#4fc3f7;font-weight:600}}
.dr-btn{{display:flex;align-items:center;gap:6px;background:#1a1d2e;color:#aaa;
         border:1px solid #1e2235;border-radius:6px;padding:5px 12px;
         font-size:.78rem;cursor:pointer;white-space:nowrap}}
.dr-btn:hover{{border-color:#4fc3f7;color:#fff}}
/* calendar modal */
.dp-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.65);
             z-index:9998;align-items:center;justify-content:center}}
.dp-modal{{background:#151824;border:1px solid #252c45;border-radius:10px;
           width:600px;max-width:96vw;box-shadow:0 12px 40px rgba(0,0,0,.6)}}
.dp-hdr{{display:flex;justify-content:space-between;align-items:center;
         padding:14px 20px 10px;border-bottom:1px solid #1e2235}}
.dp-hdr span{{font-size:.85rem;color:#ccc;font-weight:600}}
.dp-x{{background:none;border:none;color:#666;font-size:1.1rem;cursor:pointer;padding:2px 6px}}
.dp-x:hover{{color:#fff}}
.dp-cals{{display:flex;gap:0;padding:16px 20px 10px}}
.dp-cal{{flex:1}}
.dp-sep{{width:1px;background:#1e2235;margin:0 14px}}
.dp-mhdr{{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px}}
.dp-mhdr span{{font-size:.82rem;font-weight:600;color:#ccc}}
.dp-nav{{background:none;border:none;color:#666;cursor:pointer;font-size:1.4rem;
          padding:0 5px;line-height:1}}
.dp-nav:hover{{color:#4fc3f7}}
.dp-grid{{display:grid;grid-template-columns:repeat(7,1fr);gap:2px}}
.dp-dh{{text-align:center;font-size:.68rem;color:#555;padding:3px 0;font-weight:600}}
.dp-c{{text-align:center;padding:5px 2px;font-size:.78rem;cursor:pointer;
        border-radius:4px;color:#aaa;transition:.1s;user-select:none}}
.dp-c:hover{{background:#1e2235;color:#fff}}
.dp-c.dp-rng{{background:#0a2335;color:#7ec8e3;border-radius:0}}
.dp-c.dp-sel{{background:#4fc3f7 !important;color:#000 !important;
              font-weight:700;border-radius:4px !important}}
.dp-c.dp-sel-s{{border-radius:4px 0 0 4px !important}}
.dp-c.dp-sel-e{{border-radius:0 4px 4px 0 !important}}
.dp-empty{{cursor:default;pointer-events:none}}
.dp-ftr{{display:flex;align-items:center;gap:10px;padding:12px 20px;
          border-top:1px solid #1e2235}}
#dp-lbl{{flex:1;font-size:.8rem;color:#555}}
.dp-clr{{color:#4fc3f7;font-size:.78rem;cursor:pointer;background:none;border:none;
          text-decoration:underline;padding:0}}
.dp-clr:hover{{color:#fff}}
.dp-cancel{{background:#1a1d2e;color:#888;border:1px solid #1e2235;border-radius:6px;
             padding:6px 16px;cursor:pointer;font-size:.8rem}}
.dp-cancel:hover{{color:#ccc}}
.dp-apply{{background:#4fc3f7;color:#000;border:none;border-radius:6px;
            padding:6px 18px;cursor:pointer;font-size:.8rem;font-weight:700}}
.dp-apply:hover{{background:#76d7f5}}

.bus-tag{{font-family:monospace;font-weight:700;font-size:.83rem;color:#ecf0f1}}
.tid-link{{color:#4fc3f7;font-size:.72rem}}

/* summary cell */
.summary-cell{{max-width:280px;color:#aaa;font-size:.77rem;line-height:1.45;word-break:break-word}}
.assignee-cell{{font-size:.78rem;color:#bbb;white-space:nowrap}}
.ops-cell{{max-width:240px;font-size:.75rem;color:#8a9bc4;line-height:1.4;word-break:break-word}}
.ops-cell .ops-note{{border-left:2px solid #2a3a5a;padding-left:6px;margin-bottom:4px;color:#8a9bc4}}
.summary-cell .flow{{color:#bbb;margin-bottom:2px}}
.summary-cell .quote{{color:#8a9bc4;font-style:italic}}
.sort-asc::after{{content:' ↑';color:#4fc3f7}}
.sort-desc::after{{content:' ↓';color:#4fc3f7}}
.no-data{{text-align:center;padding:40px;color:#444;font-size:.9rem}}
.repeat-bus{{background:#1f1500!important;border-left:3px solid #f39c12}}
.repeat-bus:hover{{background:#2a1c00!important}}
.repeat-bus-tag{{color:#f39c12!important}}
.repeat-badge{{display:inline-block;background:#f39c12;color:#000;border-radius:4px;
               padding:0px 5px;font-size:.65rem;font-weight:800;margin-left:4px;vertical-align:middle}}
.footer{{text-align:center;padding:18px;color:#2a2e40;font-size:.72rem;border-top:1px solid #151824}}
.plot-btn{{background:#1a1d2e;color:#4fc3f7;border:1px solid #1e3a4a;border-radius:4px;
           padding:2px 8px;font-size:.72rem;cursor:pointer;white-space:nowrap}}
.plot-btn:hover{{background:#0d2a40;border-color:#4fc3f7}}
.plot-modal-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.75);z-index:9999;
                      align-items:center;justify-content:center}}
.plot-modal-overlay.open{{display:flex}}
.plot-modal{{background:#111420;border:1px solid #1e2235;border-radius:12px;width:620px;
             max-width:95vw;max-height:90vh;display:flex;flex-direction:column;overflow:hidden}}
.plot-modal-hdr{{padding:14px 18px;border-bottom:1px solid #1e2235;display:flex;
                 align-items:flex-start;gap:12px}}
.plot-modal-hdr h3{{font-size:.9rem;color:#eee;flex:1;line-height:1.4}}
.plot-close{{background:none;border:none;color:#555;font-size:1.3rem;cursor:pointer;
             padding:0 4px;line-height:1}}
.plot-close:hover{{color:#fff}}
#plot-map{{width:100%;height:380px}}
.plot-info{{padding:12px 18px;display:flex;gap:20px;flex-wrap:wrap;border-top:1px solid #1e2235}}
.plot-stat{{font-size:.78rem}}
.plot-stat .lbl{{color:#555;text-transform:uppercase;font-size:.68rem;letter-spacing:.4px}}
.plot-stat .val{{color:#ddd;font-weight:600;margin-top:2px}}
.refresh-btn{{display:flex;align-items:center;gap:6px;background:#0d2a40;border:1px solid #4fc3f7;
  color:#4fc3f7;border-radius:8px;padding:8px 16px;font-size:.82rem;font-weight:600;
  cursor:pointer;transition:background .2s;white-space:nowrap}}
.refresh-btn:hover{{background:#143d58}}
.refresh-btn.busy{{opacity:.6;cursor:not-allowed;border-color:#555;color:#888}}
.refresh-btn.busy::before{{content:'';display:inline-block;width:12px;height:12px;
  border:2px solid #555;border-top-color:#4fc3f7;border-radius:50%;
  animation:spin .8s linear infinite;margin-right:4px}}
@keyframes spin{{to{{transform:rotate(360deg)}}}}
</style>
</head>
<body>

<div class="topbar">
  <h1>Cityflo CS &mdash; <em>Missed Passenger Report</em></h1>
  <div class="top-meta">
    <strong>{start_label} &rarr; {end_label}</strong><br>
    <span id="gen-ts">Generated {generated_at} IST</span>
  </div>
  <button class="refresh-btn" id="refresh-btn" onclick="liveRefresh()">&#8635; Refresh Data</button>
</div>

<!-- KPI Cards -->
<div class="kpi-row">
  <div class="kpi k-total">
    <div class="kpi-val" id="kpi-total">{total}</div>
    <div class="kpi-label">Total Reports</div>
    <div class="kpi-sub" id="kpi-total-sub">Across all channels</div>
  </div>
  <div class="kpi k-uniq">
    <div class="kpi-val" id="kpi-uniq">{unique_trips}</div>
    <div class="kpi-label">Unique Bus Trips</div>
    <div class="kpi-sub">Distinct bus + date</div>
  </div>
  <div class="kpi k-dns">
    <div class="kpi-val" id="kpi-dns">{dns}</div>
    <div class="kpi-label">Bus Did Not Stop</div>
    <div class="kpi-sub" id="kpi-dns-pct">{pct(dns)}% of reports</div>
  </div>
  <div class="kpi k-ble">
    <div class="kpi-val" id="kpi-ble">{ble}</div>
    <div class="kpi-label">Bus Left Early</div>
    <div class="kpi-sub" id="kpi-ble-pct">{pct(ble)}% of reports</div>
  </div>
  <div class="kpi k-confirmed" onclick="filterConfirmed()" style="cursor:pointer">
    <div class="kpi-val" id="kpi-confirmed">{confirmed}</div>
    <div class="kpi-label">Confirmed by Ops</div>
    <div class="kpi-sub">Plot location verified</div>
  </div>

</div>

<!-- Charts -->
<div class="charts-row">
  <div class="chart-card main">
    <div class="chart-title">Day-wise Reports by Reason</div>
    <canvas id="dayChart" height="95"></canvas>
  </div>
  <div class="chart-card side">
    <div class="chart-title">By City</div>
    <canvas id="cityChart" height="160"></canvas>
  </div>
</div>

<!-- City-wise Breakdown -->
<div class="city-row">
  <div class="city-row-label">City Breakdown</div>
  <div class="city-cards" id="city-cards-container">{city_cards_html}</div>
</div>

<!-- Channel Strip -->
<div class="ch-strip">
  <span class="ch-label">Channel:</span>
  {ch_html}
</div>

<!-- Filter Bar -->
<div class="filter-bar">
  <label>Reason</label>
  <select id="f-reason" onchange="applyFilters()">
    <option value="">All</option>{reason_opts}
  </select>
  <label>City</label>
  <select id="f-city" onchange="applyFilters()">
    <option value="">All</option>{city_opts}
  </select>
  <label>Channel</label>
  <select id="f-channel" onchange="applyFilters()">
    <option value="">All</option>{ch_opts}
  </select>
  <label>Status</label>
  <select id="f-status" onchange="applyFilters()">
    <option value="">All</option>
    <option value="Open">Open</option>
    <option value="Resolved">Resolved</option>
  </select>
  <label>Ops</label>
  <select id="f-confirmed" onchange="applyFilters()" style="min-width:130px">
    <option value="">All</option>
    <option value="confirmed">&#10003; Confirmed</option>
    <option value="pending">&#9711; Pending from Ops</option>
  </select>
  <label>Bus&nbsp;No</label>
  <input id="f-bus" type="text" placeholder="e.g. C504" oninput="applyFilters()">
  <label>Search</label>
  <input id="f-search" type="text" placeholder="keyword in summary" oninput="applyFilters()">
  <span class="f-count" id="f-count"></span>
  <button onclick="copyFilteredIDs()"
    style="background:#1e2235;color:#4fc3f7;border:1px solid #1e3a4a;border-radius:6px;
           padding:5px 12px;cursor:pointer;font-size:.76rem">
    &#128203; Copy IDs
  </button>
  <button onclick="copyTelegramExtract()"
    style="background:#0d2a40;color:#5dade2;border:1px solid #5dade250;border-radius:6px;
           padding:5px 12px;cursor:pointer;font-size:.76rem">
    &#9992; Telegram Extract
  </button>
  <button onclick="downloadExcel()"
    style="background:#1a3a1a;color:#2ecc71;border:1px solid #27ae6050;border-radius:6px;
           padding:5px 12px;cursor:pointer;font-size:.76rem">
    &#8659; Excel
  </button>
  <button class="btn-reset" onclick="resetFilters()">Reset</button>
  <button onclick="location.reload()"
    style="background:#1a1a3a;color:#a78bfa;border:1px solid #7c3aed60;border-radius:6px;
           padding:5px 12px;cursor:pointer;font-size:.76rem" title="Reload latest data from disk">
    &#8635; Refresh
  </button>
  {refresh_badge}
</div>
<div class="filter-bar" style="padding-top:8px;padding-bottom:8px">
  <label>Date</label>
  <button class="dr-btn" onclick="openDatePicker()">&#128197; <span id="dr-lbl" style="color:#555">All dates</span></button>
  <div class="week-picker-wrap">
    {week_btns_html}
  </div>
</div>

<div class="sec-hdr">All Records</div>
<div class="tbl-wrap">
  <table>
    <thead>
      <tr>
        <th onclick="sortTable(0)">Date</th>
        <th onclick="sortTable(1)">Bus No</th>
        <th onclick="sortTable(2)">City</th>
        <th onclick="sortTable(3)">Reason</th>
        <th onclick="sortTable(4)">Channel</th>
        <th onclick="sortTable(5)">Status</th>
        <th>Customer Interaction</th>
        <th>CS Note</th>
        <th>Ops Comments</th>
        <th onclick="sortTable(9)">Assigned To</th>
        <th onclick="sortTable(10)">Ops Assignee</th>
        <th onclick="sortTable(11)">Ticket ID</th>
      </tr>
    </thead>
    <tbody id="tbl-body">__STATIC_ROWS__
    <tr id="no-data-row" style="display:none"><td colspan="12" class="no-data">No records match your filters.</td></tr>
    </tbody>
  </table>
</div>

<div class="footer">
  Cityflo CS Quality &bull; Missed Pax Dashboard &bull; Data from PostgreSQL (read-only)
</div>

<!-- Date Range Picker Modal -->
<div class="dp-overlay" id="dp-overlay" onclick="if(event.target===this)closeDatePicker()">
  <div class="dp-modal">
    <div class="dp-hdr">
      <span>Select Date Range</span>
      <button class="dp-x" onclick="closeDatePicker()">&#10005;</button>
    </div>
    <div class="dp-cals">
      <div id="dp-cal-l" class="dp-cal"></div>
      <div class="dp-sep"></div>
      <div id="dp-cal-r" class="dp-cal"></div>
    </div>
    <div class="dp-ftr">
      <span id="dp-lbl">Click a date to start</span>
      <button class="dp-clr" onclick="clearDateRange()">Clear</button>
      <button class="dp-cancel" onclick="closeDatePicker()">Cancel</button>
      <button class="dp-apply" onclick="applyDateRange()">Apply</button>
    </div>
  </div>
</div>

<!-- GPS Plot Modal -->
<div class="plot-modal-overlay" id="plot-overlay" onclick="if(event.target===this)closePlot()">
  <div class="plot-modal">
    <div class="plot-modal-hdr">
      <h3 id="plot-title">Bus GPS Plot</h3>
      <button class="plot-close" onclick="closePlot()">&#10005;</button>
    </div>
    <div id="plot-map"></div>
    <div class="plot-info" id="plot-info"></div>
  </div>
</div>

<script>
const ROWS = {rows_json};
const ALL_DATES = {chart_labels};
let filtered = [...ROWS];
let sortCol = 0, sortDir = 1;

const PLOT_MAP = {{}};
ROWS.forEach(r => {{ if(r.plot) PLOT_MAP[r.ticket_id] = {{...r.plot, bus_num: r.bus_num}}; }});

const REASON_B = {{'Bus Did Not Stop':'b-dns','Bus Left Early':'b-ble'}};
const CITY_B   = {{'Mumbai':'b-Mumbai','Delhi':'b-Delhi','Kolkata':'b-Kolkata',
                    'Hyderabad':'b-Hyderabad','Unknown':'b-Unknown'}};
const CHAN_B   = {{'Sage':'b-Sage','Ticket':'b-Ticket'}};

function rsLabel(r){{
  if(r==='Bus Did Not Stop') return 'Did Not Stop';
  if(r==='Bus Left Early') return 'Left Early';
  return r;
}}

function escHtml(s){{
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}

function summaryHTML(s){{
  if(!s||s==='—') return '<span style="color:#444">—</span>';
  const parts = s.split(' | ');
  let html='';
  for(const part of parts){{
    const e = escHtml(part);
    if(part.startsWith('"')&&part.endsWith('"'))
      html+=`<div class="quote">${{e}}</div>`;
    else
      html+=`<div class="flow">${{e}}</div>`;
  }}
  return html;
}}

/* Table rows are rendered by Python as static HTML with data-* attributes.
   JS only shows/hides them — no innerHTML replacement needed. */
function updateRowCount(){{
  const busDay = {{}};
  ROWS.forEach(r=>{{ if(r.bus_num!=='Unknown') busDay[r.bus_num+'|'+r.date_iso]=(busDay[r.bus_num+'|'+r.date_iso]||0)+1; }});
  const uniq = new Set(filtered.filter(r=>r.bus_num!=='Unknown').map(r=>r.bus_num+'|'+r.date_iso)).size;
  const repeatTrips = new Set(filtered.filter(r=>r.bus_num!=='Unknown'&&(busDay[r.bus_num+'|'+r.date_iso]||0)>1).map(r=>r.bus_num+'|'+r.date_iso)).size;
  document.getElementById('f-count').innerHTML =
    `<strong>${{filtered.length}}</strong> records &nbsp;|&nbsp; <strong>${{uniq}}</strong> unique trips` +
    (repeatTrips ? ` &nbsp;|&nbsp; <strong style="color:#f39c12">${{repeatTrips}}</strong> <span style="color:#f39c12">repeat bus-days</span>` : '');
  const noData = document.getElementById('no-data-row');
  if(noData) noData.style.display = filtered.length ? 'none' : '';
}}

// ── Date range filter ────────────────────────────────────────────────────
let filterDateStart = null, filterDateEnd = null;

function _fmtIso(iso){{
  const d=new Date(iso+'T00:00:00');
  return d.getDate()+' '+['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][d.getMonth()];
}}

function _setDrLabel(){{
  const el=document.getElementById('dr-lbl');
  if(!el) return;
  if(filterDateStart && filterDateEnd){{
    el.textContent = filterDateStart===filterDateEnd ? _fmtIso(filterDateStart)
      : _fmtIso(filterDateStart)+' – '+_fmtIso(filterDateEnd);
    el.style.color='#4fc3f7';
  }} else {{ el.textContent='All dates'; el.style.color='#555'; }}
}}

function selectWeek(btn){{
  // Toggle: clicking active week clears it
  if(btn.classList.contains('active')){{
    btn.classList.remove('active');
    filterDateStart=filterDateEnd=null;
    _setDrLabel(); applyFilters(); return;
  }}
  document.querySelectorAll('.week-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  filterDateStart=btn.dataset.start; filterDateEnd=btn.dataset.end;
  _setDrLabel(); applyFilters();
}}

// ── Calendar picker ───────────────────────────────────────────────────────
let dpS=null,dpE=null,dpPickE=false,dpY,dpM;
const _DPM=['January','February','March','April','May','June',
            'July','August','September','October','November','December'];
const _DPD=['Su','Mo','Tu','We','Th','Fr','Sa'];

function openDatePicker(){{
  const now=new Date();
  if(filterDateStart){{
    const d=new Date(filterDateStart+'T00:00:00'); dpY=d.getFullYear(); dpM=d.getMonth();
    dpS=filterDateStart; dpE=filterDateEnd;
  }} else {{ dpY=now.getFullYear(); dpM=now.getMonth()-1; if(dpM<0){{dpM=11;dpY--;}} dpS=dpE=null; }}
  dpPickE=false; _dpRender();
  document.getElementById('dp-overlay').style.display='flex';
}}
function closeDatePicker(){{ document.getElementById('dp-overlay').style.display='none'; }}
function dpNav(n){{ dpM+=n; if(dpM>11){{dpM=0;dpY++;}} if(dpM<0){{dpM=11;dpY--;}} _dpRender(); }}

function _dpRender(){{
  let rM=dpM+1,rY=dpY; if(rM>11){{rM=0;rY++;}}
  _dpMonth('dp-cal-l',dpY,dpM,true);
  _dpMonth('dp-cal-r',rY,rM,false);
  const lbl=document.getElementById('dp-lbl');
  if(lbl) lbl.textContent = dpS&&dpE ? _fmtIso(dpS)+(dpS!==dpE?' – '+_fmtIso(dpE):'')
    : dpS ? _fmtIso(dpS)+' – ?' : 'Click a start date';
}}

function _dpMonth(id,year,month,showPrev){{
  const el=document.getElementById(id); if(!el) return;
  const fd=new Date(year,month,1).getDay();
  const dim=new Date(year,month+1,0).getDate();
  let h=`<div class="dp-mhdr">`;
  h+=showPrev?`<button class="dp-nav" onclick="dpNav(-1)">&#8249;</button>`:`<div></div>`;
  h+=`<span>${{_DPM[month]}} ${{year}}</span>`;
  h+=!showPrev?`<button class="dp-nav" onclick="dpNav(1)">&#8250;</button>`:`<div></div>`;
  h+=`</div><div class="dp-grid">`;
  _DPD.forEach(d=>h+=`<div class="dp-dh">${{d}}</div>`);
  for(let i=0;i<fd;i++) h+=`<div class="dp-c dp-empty"></div>`;
  for(let d=1;d<=dim;d++){{
    const iso=`${{year}}-${{String(month+1).padStart(2,'0')}}-${{String(d).padStart(2,'0')}}`;
    let cls='dp-c';
    if(dpS&&dpE&&iso>dpS&&iso<dpE) cls+=' dp-rng';
    if(iso===dpS) cls+=(dpS!==dpE?' dp-sel dp-sel-s':' dp-sel');
    if(iso===dpE&&dpE!==dpS) cls+=' dp-sel dp-sel-e';
    h+=`<div class="${{cls}}" onclick="dpClick('${{iso}}')">${{d}}</div>`;
  }}
  h+=`</div>`;
  el.innerHTML=h;
}}

function dpClick(iso){{
  if(!dpPickE){{ dpS=iso; dpE=null; dpPickE=true; }}
  else {{ if(iso<dpS){{dpE=dpS;dpS=iso;}}else{{dpE=iso;}} dpPickE=false; }}
  _dpRender();
}}

function applyDateRange(){{
  if(dpS){{
    filterDateStart=dpS; filterDateEnd=dpE||dpS;
    document.querySelectorAll('.week-btn').forEach(b=>
      b.classList.toggle('active',b.dataset.start===filterDateStart&&b.dataset.end===filterDateEnd));
    _setDrLabel(); applyFilters();
  }}
  closeDatePicker();
}}

function clearDateRange(){{
  dpS=dpE=null; dpPickE=false;
  filterDateStart=filterDateEnd=null;
  document.querySelectorAll('.week-btn').forEach(b=>b.classList.remove('active'));
  _setDrLabel(); _dpRender(); applyFilters(); closeDatePicker();
}}

function distBadge(minDist, dwell){{
  let color, label;
  if(dwell !== null && dwell !== undefined && dwell < 15){{
    color='#e74c3c'; label='Pass-through';
  }} else if(minDist <= 50){{
    color='#00b894'; label='At stop';
  }} else if(minDist <= 150){{
    color='#f39c12'; label=minDist+'m offset';
  }} else if(minDist <= 300){{
    color='#e67e22'; label=minDist+'m ahead';
  }} else {{
    color='#e74c3c'; label=minDist+'m away';
  }}
  return '<div style="margin-top:4px"><span style="background:'+color+'22;color:'+color+
    ';border:1px solid '+color+'55;border-radius:4px;padding:1px 6px;font-size:.68rem;font-weight:700">'+
    '&#128205; '+label+'</span></div>';
}}

function opsHTML(notes, teamLabel, badgeColor){{
  if(!notes||!notes.length) return '<span style="color:#333">—</span>';
  return notes.map(n=>{{
    const text = escHtml(typeof n === 'string' ? n : (n.text||''));
    const name = escHtml(typeof n === 'object' ? (n.name||'') : '');
    const hdr = name
      ? `<div style="margin-bottom:3px"><span style="background:${{badgeColor}};color:#fff;font-size:.62rem;font-weight:700;padding:1px 6px;border-radius:3px;letter-spacing:.4px;text-transform:uppercase">${{teamLabel}}</span> <span style="color:#aaa;font-size:.72rem">${{name}}</span></div>`
      : '';
    return `<div class="ops-note">${{hdr}}${{text}}</div>`;
  }}).join('');
}}

function applyFilters(){{
  const reason  = document.getElementById('f-reason').value;
  const city    = document.getElementById('f-city').value;
  const channel = document.getElementById('f-channel').value;
  const status  = document.getElementById('f-status').value;
  const bus     = document.getElementById('f-bus').value.trim().toUpperCase();
  const search  = document.getElementById('f-search').value.trim().toLowerCase();
  const confVal = document.getElementById('f-confirmed').value;

  const allTrs = Array.from(document.querySelectorAll('#tbl-body tr[data-tid]'));
  const visibleTids = new Set();
  allTrs.forEach(tr => {{
    const d = tr.dataset;
    const show =
      (!reason  || d.reason  === reason)  &&
      (!city    || d.city    === city)    &&
      (!channel || d.channel === channel) &&
      (!status  || d.status  === status)  &&
      (confVal==='' || (confVal==='confirmed' && d.confirmed==='true') ||
                       (confVal==='pending'   && d.confirmed==='false')) &&
      (!filterDateStart || (d.date >= filterDateStart && d.date <= filterDateEnd)) &&
      (!bus    || (d.bus||'').toUpperCase().includes(bus)) &&
      (!search || (d.search||'').includes(search));
    tr.style.display = show ? '' : 'none';
    if(show) visibleTids.add(+d.tid);
  }});
  filtered = ROWS.filter(r => visibleTids.has(r.ticket_id));
  updateRowCount();
  updateDashboard();
}}

function resetFilters(){{
  ['f-reason','f-city','f-channel','f-status','f-confirmed'].forEach(id=>document.getElementById(id).value='');
  ['f-bus','f-search'].forEach(id=>document.getElementById(id).value='');
  filterDateStart=filterDateEnd=null;
  document.querySelectorAll('.week-btn').forEach(b=>b.classList.remove('active'));
  _setDrLabel();
  applyFilters();
}}

function filterConfirmed(){{
  document.getElementById('f-confirmed').value = 'confirmed';
  applyFilters();
  document.getElementById('tbl-body').scrollIntoView({{behavior:'smooth'}});
}}

function liveRefresh(){{
  const btn = document.getElementById('refresh-btn');
  if(btn.classList.contains('busy')) return;
  btn.classList.add('busy');
  btn.textContent = ' Refreshing...';
  fetch('/api/refresh', {{method:'POST'}})
    .then(r => r.json())
    .then(data => {{
      if(data.status === 'ok'){{
        location.reload();
      }} else {{
        btn.classList.remove('busy');
        btn.textContent = '↻ Refresh Data';
        alert('Refresh failed: ' + (data.msg || data.status));
      }}
    }})
    .catch(err => {{
      btn.classList.remove('busy');
      btn.textContent = '↻ Refresh Data';
      alert('Server not reachable. Run dashboard_server.py to enable live refresh.');
    }});
}}

function filterCity(city){{
  document.getElementById('f-city').value = city;
  applyFilters();
  document.getElementById('tbl-body').scrollIntoView({{behavior:'smooth'}});
}}

function sortTable(col){{
  const ths=document.querySelectorAll('thead th');
  if(sortCol===col) sortDir*=-1; else {{sortCol=col;sortDir=1;}}
  ths.forEach((th,i)=>th.className=i===col?(sortDir===1?'sort-asc':'sort-desc'):'');
  applyFilters();
  sortAndRender();
}}

function sortAndRender(){{
  const domKeys = ['date','bus','city','reason','channel','status','summary','','','assigned','opsassigned','tid'];
  const dk = domKeys[sortCol] || 'date';

  const tb = document.getElementById('tbl-body');
  // Build rowMap once — used in sort comparator (avoids repeated DOM queries)
  const rowMap = {{}};
  tb.querySelectorAll('tr[data-tid]').forEach(tr => {{ rowMap[+tr.dataset.tid] = tr; }});

  // Sort visible (filtered) rows by chosen column
  const sortedTids = filtered.map(r => r.ticket_id).slice().sort((a, b) => {{
    const av = (rowMap[a] && dk ? rowMap[a].dataset[dk] : '') || '';
    const bv = (rowMap[b] && dk ? rowMap[b].dataset[dk] : '') || '';
    return av.localeCompare(bv, undefined, {{numeric:true}}) * sortDir;
  }});

  // Append sorted visible rows to end of tbody (hidden rows stay, display:none)
  sortedTids.forEach(tid => {{ if(rowMap[tid]) tb.appendChild(rowMap[tid]); }});

  updateRowCount();
  updateDashboard();
}}

/* Charts — wrapped in try/catch so a CDN failure never blocks the table */
try {{
  window.dayChart = new Chart(document.getElementById('dayChart').getContext('2d'),{{
    type:'bar',
    data:{{
      labels:{chart_labels},
      datasets:[
        {{label:'Bus Did Not Stop',data:{chart_dns},backgroundColor:'rgba(231,76,60,0.85)',stack:'s'}},
        {{label:'Bus Left Early',  data:{chart_ble},backgroundColor:'rgba(243,156,18,0.85)',stack:'s'}},
      ]
    }},
    options:{{
      responsive:true,
      plugins:{{legend:{{labels:{{color:'#888',font:{{size:11}}}}}}}},
      scales:{{
        x:{{stacked:true,ticks:{{color:'#666'}},grid:{{color:'#1a1d2e'}}}},
        y:{{stacked:true,ticks:{{color:'#666'}},grid:{{color:'#1a1d2e'}},beginAtZero:true}}
      }}
    }}
  }});
  window.cityChart = new Chart(document.getElementById('cityChart').getContext('2d'),{{
    type:'doughnut',
    data:{{
      labels:{donut_labels},
      datasets:[{{data:{donut_data},backgroundColor:{donut_colors},borderWidth:0}}]
    }},
    options:{{
      responsive:true,
      plugins:{{
        legend:{{position:'bottom',labels:{{color:'#888',font:{{size:11}},padding:10}}}}
      }}
    }}
  }});
}} catch(e) {{ console.warn('Charts unavailable:', e); }}

function updateDashboard(){{
  const total     = filtered.length;
  const dns       = filtered.filter(r=>r.reason==='Bus Did Not Stop').length;
  const ble       = filtered.filter(r=>r.reason==='Bus Left Early').length;
  const conf      = filtered.filter(r=>r.confirmed).length;
  const uniqueTrips = new Set(filtered.filter(r=>r.bus_num!=='Unknown').map(r=>r.bus_num+'|'+r.date_iso)).size;
  function pct(n){{ return total ? Math.round(n/total*100) : 0; }}
  const el = id => document.getElementById(id);
  if(el('kpi-total'))     el('kpi-total').textContent     = total;
  if(el('kpi-uniq'))      el('kpi-uniq').textContent      = uniqueTrips;
  if(el('kpi-dns'))       el('kpi-dns').textContent       = dns;
  if(el('kpi-dns-pct'))   el('kpi-dns-pct').textContent   = pct(dns)+'% of reports';
  if(el('kpi-ble'))       el('kpi-ble').textContent       = ble;
  if(el('kpi-ble-pct'))   el('kpi-ble-pct').textContent   = pct(ble)+'% of reports';
  if(el('kpi-confirmed')) el('kpi-confirmed').textContent = conf;
  // City breakdown cards
  const cityOrder = ['Mumbai','Delhi','Kolkata','Hyderabad','Unknown'];
  const cityColor = {{Mumbai:'#3498db',Delhi:'#e74c3c',Kolkata:'#2ecc71',Hyderabad:'#f39c12',Unknown:'#636e72'}};
  const cs = {{}};
  filtered.forEach(r=>{{
    if(!cs[r.city]) cs[r.city]={{total:0,uniq:new Set()}};
    cs[r.city].total++;
    if(r.bus_num!=='Unknown') cs[r.city].uniq.add(r.bus_num+'|'+r.date_iso);
  }});
  const ctn = document.getElementById('city-cards-container');
  if(ctn){{
    ctn.innerHTML = cityOrder.filter(c=>cs[c]).map(c=>{{
      const col = cityColor[c]||'#aaa';
      return `<div class="city-card" onclick="filterCity('${{c}}')" title="Click to filter by ${{c}}">` +
        `<div class="city-card-name" style="border-left:3px solid ${{col}};padding-left:8px">${{c}}</div>` +
        `<div class="city-card-nums">` +
        `<span class="city-stat"><span class="city-stat-val">${{cs[c].total}}</span><span class="city-stat-lbl">Total</span></span>` +
        `<span class="city-stat"><span class="city-stat-val">${{cs[c].uniq.size}}</span><span class="city-stat-lbl">Unique</span></span>` +
        `</div></div>`;
    }}).join('');
  }}
  // Charts
  try{{
    if(window.dayChart){{
      window.dayChart.data.datasets[0].data = ALL_DATES.map(d=>filtered.filter(r=>r.date_iso===d&&r.reason==='Bus Did Not Stop').length);
      window.dayChart.data.datasets[1].data = ALL_DATES.map(d=>filtered.filter(r=>r.date_iso===d&&r.reason==='Bus Left Early').length);
      window.dayChart.update('none');
    }}
    if(window.cityChart){{
      const active = cityOrder.filter(c=>cs[c]);
      window.cityChart.data.labels = active;
      window.cityChart.data.datasets[0].data = active.map(c=>cs[c].total);
      window.cityChart.data.datasets[0].backgroundColor = active.map(c=>cityColor[c]||'#aaa');
      window.cityChart.update('none');
    }}
  }} catch(e){{ console.warn('Chart update failed:',e); }}
}}

applyFilters();

function copyToClipboard(text, btn, resetHTML){{
  /* Works on file:// (no HTTPS required) */
  try {{
    if(navigator.clipboard && window.isSecureContext){{
      navigator.clipboard.writeText(text).then(()=>{{
        btn.textContent='✓ Copied!';
        setTimeout(()=>{{btn.innerHTML=resetHTML;}},2000);
      }});
    }} else {{
      const ta=document.createElement('textarea');
      ta.value=text; ta.style.position='fixed'; ta.style.opacity='0';
      document.body.appendChild(ta); ta.focus(); ta.select();
      document.execCommand('copy');
      document.body.removeChild(ta);
      btn.textContent='✓ Copied!';
      setTimeout(()=>{{btn.innerHTML=resetHTML;}},2000);
    }}
  }} catch(e){{
    prompt('Copy this text (Ctrl+A then Ctrl+C):',text);
  }}
}}

/* ── GPS Plot ── */
let _plotMap = null;

function openPlot(ticketId){{
  const plot = PLOT_MAP[ticketId];
  if(!plot) return;
  const busNum = plot.bus_num || '';
  const overlay = document.getElementById('plot-overlay');
  document.getElementById('plot-title').textContent =
    'Bus ' + busNum + ' · Stop: ' + (plot.stop_name || 'Unknown') + ' · #' + ticketId;

  // Distance verdict
  const minDist = plot.min_dist;
  let distVerdict, distColor;
  if(minDist === null || minDist === undefined){{
    distVerdict = 'No GPS data'; distColor = '#888';
  }} else if(minDist <= 50){{
    distVerdict = 'At stop (' + minDist + 'm) — bus was at designated location'; distColor = '#00b894';
  }} else if(minDist <= 150){{
    distVerdict = 'Near stop (' + minDist + 'm) — within geofence'; distColor = '#f39c12';
  }} else if(minDist <= 300){{
    distVerdict = 'Stopped offset (' + minDist + 'm away) — halted ahead/behind'; distColor = '#e67e22';
  }} else {{
    distVerdict = 'Far from stop (' + minDist + 'm) — genuine DNS'; distColor = '#e74c3c';
  }}

  // Dwell verdict
  const dwell = plot.dwell_sec;
  let dwellVerdict, dwellColor;
  if(dwell === null || dwell === undefined){{
    dwellVerdict = '—'; dwellColor = '#888';
  }} else if(dwell < 15){{
    dwellVerdict = 'Pass-through (' + dwell + 's)'; dwellColor = '#e74c3c';
  }} else if(dwell < 60){{
    dwellVerdict = 'Bus Halted (' + dwell + 's)'; dwellColor = '#00b894';
  }} else {{
    dwellVerdict = 'Bus Halted (' + dwell + 's)'; dwellColor = '#00b894';
  }}

  // Info bar
  const info = document.getElementById('plot-info');
  info.innerHTML =
    '<div class="plot-stat" style="flex:2;border-right:1px solid #1e2235;padding-right:16px;margin-right:4px">' +
    '<div class="lbl">Location Verdict</div>' +
    '<div class="val" style="color:' + distColor + ';font-size:.9rem">' + distVerdict + '</div></div>' +
    '<div class="plot-stat"><div class="lbl">Dwell</div>' +
    '<div class="val" style="color:' + dwellColor + '">' + dwellVerdict + '</div></div>' +
    '<div class="plot-stat"><div class="lbl">Arrival GPS → Stop</div><div class="val">' + (plot.arr_dist !== null && plot.arr_dist !== undefined ? plot.arr_dist + 'm' : '—') + '</div></div>' +
    '<div class="plot-stat"><div class="lbl">Departure GPS → Stop</div><div class="val">' + (plot.dep_dist !== null && plot.dep_dist !== undefined ? plot.dep_dist + 'm' : '—') + '</div></div>' +
    '<div class="plot-stat"><div class="lbl">Arrival IST</div><div class="val">' + (plot.arr_ts||'—') + '</div></div>' +
    '<div class="plot-stat"><div class="lbl">Departure IST</div><div class="val">' + (plot.dep_ts||'—') + '</div></div>';

  overlay.classList.add('open');

  // Build map
  setTimeout(()=>{{
    if(_plotMap){{ _plotMap.remove(); _plotMap=null; }}
    const center = plot.stop_lat && plot.stop_lng ? [plot.stop_lat, plot.stop_lng]
                 : plot.arr_lat  ? [plot.arr_lat, plot.arr_lng] : [0,0];
    _plotMap = L.map('plot-map').setView(center, 17);
    L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png',{{
      attribution:'© OpenStreetMap',maxZoom:19
    }}).addTo(_plotMap);

    // Stop location (star)
    if(plot.stop_lat && plot.stop_lng){{
      L.circle([plot.stop_lat, plot.stop_lng], {{
        radius:150, color:'#4fc3f7', fillColor:'#4fc3f7', fillOpacity:.08, weight:2
      }}).addTo(_plotMap).bindPopup('Stop geofence (150m)');
      L.marker([plot.stop_lat, plot.stop_lng], {{
        icon: L.divIcon({{className:'',html:'<div style="background:#4fc3f7;width:14px;height:14px;border-radius:50%;border:3px solid #fff;box-shadow:0 0 6px #4fc3f7"></div>',iconSize:[14,14],iconAnchor:[7,7]}})
      }}).addTo(_plotMap).bindPopup('★ Stop: ' + (plot.stop_name||''));
    }}

    // Arrival (green)
    if(plot.arr_lat){{
      L.marker([plot.arr_lat, plot.arr_lng], {{
        icon: L.divIcon({{className:'',html:'<div style="background:#00b894;width:12px;height:12px;border-radius:50%;border:2px solid #fff"></div>',iconSize:[12,12],iconAnchor:[6,6]}})
      }}).addTo(_plotMap).bindPopup('Arrival GPS ' + (plot.arr_ts||''));
    }}

    // Departure (orange/red)
    if(plot.dep_lat){{
      L.marker([plot.dep_lat, plot.dep_lng], {{
        icon: L.divIcon({{className:'',html:'<div style="background:#e74c3c;width:12px;height:12px;border-radius:50%;border:2px solid #fff"></div>',iconSize:[12,12],iconAnchor:[6,6]}})
      }}).addTo(_plotMap).bindPopup('Departure GPS ' + (plot.dep_ts||''));
    }}

    // Line between arrival and departure
    if(plot.arr_lat && plot.dep_lat){{
      L.polyline([[plot.arr_lat, plot.arr_lng],[plot.dep_lat, plot.dep_lng]],
        {{color:'#f39c12',weight:2,dashArray:'5,5'}}).addTo(_plotMap);
    }}
  }}, 50);
}}

function closePlot(){{
  document.getElementById('plot-overlay').classList.remove('open');
  if(_plotMap){{ _plotMap.remove(); _plotMap=null; }}
}}

function downloadExcel(){{
  if(!filtered.length){{ alert('No records to export.'); return; }}
  const confVal = document.getElementById('f-confirmed').value;
  const sheetLabel = confVal==='confirmed' ? 'Confirmed'
                   : confVal==='pending'   ? 'Pending'
                   : 'All Records';

  function gpsVerdict(plot){{
    if(!plot || plot.min_dist===null || plot.min_dist===undefined) return 'No GPS';
    if(plot.dwell_sec!==null && plot.dwell_sec!==undefined && plot.dwell_sec<15) return 'Pass-through';
    if(plot.dwell_sec!==null && plot.dwell_sec!==undefined && plot.dwell_sec>=15) return 'Bus Halted';
    if(plot.min_dist<=50)   return 'At stop';
    if(plot.min_dist<=150)  return plot.min_dist+'m offset';
    if(plot.min_dist<=300)  return plot.min_dist+'m ahead';
    return plot.min_dist+'m away';
  }}

  const rows = filtered.map(r => ({{
    'Date':             r.date_str,
    'Bus No':           r.bus_num,
    'City':             r.city,
    'Reason':           r.reason,
    'Channel':          r.channel,
    'Ticket Status':    r.status,
    'Ops Status':       r.confirmed ? 'Confirmed' : 'Pending from Ops',
    'Customer Summary': r.summary || '',
    'CS Note':          (r.cs_notes||[]).map(n=>typeof n==='string'?n:(n.text||'')).join(' | '),
    'Ops Comments':     (r.ops_notes||[]).map(n=>typeof n==='string'?n:(n.text||'')).join(' | '),
    'Assigned To':      r.assigned || '',
    'Ops Assignee':     r.ops_assignee || '',
    'GPS Verdict':      gpsVerdict(r.plot),
    'Dwell (sec)':      r.plot ? (r.plot.dwell_sec !== null && r.plot.dwell_sec !== undefined ? r.plot.dwell_sec : '') : '',
    'Dist to Stop (m)': r.plot ? (r.plot.min_dist !== null && r.plot.min_dist !== undefined ? r.plot.min_dist : '') : '',
    'Stop Name':        r.plot ? (r.plot.stop_name||'') : '',
    'Ticket ID':        r.ticket_id,
  }}));

  const ws = XLSX.utils.json_to_sheet(rows);

  // Column widths
  ws['!cols'] = [
    {{wch:12}},{{wch:10}},{{wch:11}},{{wch:18}},{{wch:9}},{{wch:12}},{{wch:18}},
    {{wch:45}},{{wch:40}},{{wch:40}},{{wch:18}},{{wch:18}},{{wch:16}},{{wch:10}},{{wch:14}},{{wch:22}},{{wch:10}}
  ];

  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, sheetLabel);

  const dateRange = '{start_label}_{end_label}'.replace(/ /g,'');
  XLSX.writeFile(wb, 'MissedPax_' + dateRange + '_' + sheetLabel + '.xlsx');
}}

function copyFilteredIDs(){{
  const ids = filtered.map(r => r.ticket_id).join(', ');
  copyToClipboard(ids, event.currentTarget, '&#128203; Copy IDs');
}}

function copyTelegramExtract(){{
  const dateRange = '{start_label} – {end_label}';
  const confVal = document.getElementById('f-confirmed').value;
  const label = confVal === 'confirmed' ? 'Confirmed Missed Passenger Report'
              : confVal === 'pending'   ? 'Pending Ops Review — Missed Passenger Report'
              :                           'Missed Passenger Report';
  let lines = ['*' + label + '*', '_' + dateRange + '_', ''];
  const byDate = {{}};
  filtered.forEach(r => {{
    if(!byDate[r.date_str]) byDate[r.date_str] = [];
    byDate[r.date_str].push(r);
  }});
  Object.keys(byDate).sort().forEach(date => {{
    lines.push('*' + date + '*');
    byDate[date].forEach(r => {{
      const _n0 = r.ops_notes && r.ops_notes.length ? r.ops_notes[0] : null;
      const note = _n0 ? (typeof _n0==='string' ? _n0 : (_n0.text||'')).slice(0,120) : '—';
      lines.push('• Bus ' + r.bus_num + ' | ' + r.city + ' | ' + r.reason + ' | #' + r.ticket_id);
      if(r.confirmed) lines.push('  _Ops: ' + note + '_');
    }});
    lines.push('');
  }});
  lines.push('Total: ' + filtered.length);
  const text = lines.join('\\n');
  copyToClipboard(text, event.currentTarget, '&#9992; Telegram Extract');
}}
{auto_reload_js}
</script>
</body>
</html>"""
    return html.replace('__STATIC_ROWS__', static_rows, 1)


# ── main ──────────────────────────────────────────────────────────────────────

def run_once(db_url, start_utc, end_utc, start_label, end_label, refresh_min=0):
    now_ist = datetime.now(IST).strftime('%d %b %Y %I:%M %p')
    print(f'[{now_ist}] Fetching data ...')
    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Missed Passenger Data.xlsx')
    excel_supplement = load_excel_supplement(excel_path, start_utc.astimezone(IST).date())
    rows = fetch_data(db_url, start_utc, end_utc, excel_supplement=excel_supplement)
    by_reason = defaultdict(int)
    for r in rows: by_reason[r['reason']] += 1
    confirmed_count = sum(1 for r in rows if r.get('confirmed'))
    print(f'  {len(rows)} records  |  DNS={by_reason["Bus Did Not Stop"]}  '
          f'BLE={by_reason["Bus Left Early"]}  |  Confirmed={confirmed_count}')
    html = build_html(rows, start_label, end_label, now_ist, refresh_min)
    with open(OUT_FILE, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'  Written → {OUT_FILE}')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--from', dest='from_date', default='2026-03-02')
    ap.add_argument('--to',   dest='to_date',
                    default=datetime.now(IST).strftime('%Y-%m-%d'))
    ap.add_argument('--watch', type=int, default=0, metavar='MINUTES')
    args = ap.parse_args()

    env    = load_env()
    db_url = env.get('DATABASE_URL')
    if not db_url:
        print('ERROR: DATABASE_URL not in env'); return

    start_dt  = datetime.strptime(args.from_date, '%Y-%m-%d').replace(tzinfo=IST)
    end_dt    = datetime.strptime(args.to_date, '%Y-%m-%d').replace(
                    hour=23, minute=59, second=59, tzinfo=IST)
    start_utc = start_dt.astimezone(timezone.utc)
    end_utc   = end_dt.astimezone(timezone.utc)

    run_once(db_url, start_utc, end_utc,
             start_dt.strftime('%d %b %Y'), end_dt.strftime('%d %b %Y'),
             refresh_min=args.watch)
    os.startfile(OUT_FILE)

    if args.watch:
        print(f'\nAuto-refresh every {args.watch} min — Ctrl+C to stop.')
        def loop():
            while True:
                time.sleep(args.watch * 60)
                try: run_once(db_url, start_utc, end_utc,
                              start_dt.strftime('%d %b %Y'), end_dt.strftime('%d %b %Y'),
                              refresh_min=args.watch)
                except Exception as e: print(f'  ERROR: {e}')
        threading.Thread(target=loop, daemon=True).start()
        try:
            while True: time.sleep(60)
        except KeyboardInterrupt:
            print('\nStopped.')


if __name__ == '__main__':
    main()
